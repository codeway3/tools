import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

DIRNAME = './tmp/2017/'
if not os.path.exists(DIRNAME):
    os.makedirs(DIRNAME)

left, right, top, bottom = [Side(style='thin', color='000000')]*4  # 设置单元格边框属性
border = Border(left=left, right=right, top=top, bottom=bottom)   # 设置单元格边框格式

deal = load_workbook('./tmp/信息工程部采购合同.xlsx', read_only=True)
model = load_workbook('./tmp/项目采购内部意见表.xlsx')
sheet = deal.active
for i, row in enumerate(tuple(sheet.rows)):
    if row[11].value == '缺评审表':
        num = row[2].value
        company_name = row[3].value
        amount = row[4].value
        content = row[5].value
        project_num = row[7].value
        project_name = row[8].value
        tmp = model
        tmp_sheet = tmp.active
        tmp_sheet['C3'].value = project_name
        tmp_sheet['F3'].value = project_num
        tmp_sheet['C4'].value = content
        tmp_sheet['C6'].value = '本次采购跟3方供应商进行询价比较，分别为{}、、。此3方供应商均有多年产品服务销售经验且在业界积累了良好的信誉度，所售产品服务保证能够达到项目需求且货期能够满足项目要求。'.format(company_name)
        tmp_sheet['C8'].value = '通过3家供应商的价格对比，综合价格、货期、供应商品质等因素，{}跟我公司合作多年，有一定的合作基础，且货期合适，价格最低，因此确定{}为最终供应商。'.format(company_name, company_name)
        tmp_sheet['C11'].value = company_name
        tmp_sheet['F11'].value = '￥{}元'.format(amount)
        for i in range(3, 18):
            for col in 'BCDEFG':
                tmp_sheet[col+str(i)].border = border
        filename = '{}{}{}.xlsx'.format(num, company_name, amount)
        tmp.save(DIRNAME+filename)
